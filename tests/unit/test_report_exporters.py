"""Tests for the four report exporters, with and without rows."""

import csv
import io
import json
from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from app.models.query import SerializedValue
from app.reporting.exporters.base import ReportData, ReportExporter
from app.reporting.exporters.csv_exporter import CsvReportExporter
from app.reporting.exporters.html_exporter import HtmlReportExporter
from app.reporting.exporters.json_exporter import JsonReportExporter
from app.reporting.exporters.pdf_exporter import PdfReportExporter
from app.reporting.exporters.xlsx_exporter import XlsxReportExporter

GENERATED_AT = datetime(2026, 7, 16, 12, 0, tzinfo=UTC)


def _report_data(rows: tuple[tuple[SerializedValue, ...], ...]) -> ReportData:
    return ReportData(
        columns=("id", "nombre"),
        rows=rows,
        title="Reporte: ventas del mes pasado",
        generated_at=GENERATED_AT,
        period_label="1 de junio de 2026 al 30 de junio de 2026",
        applied_filters=("Periodo: 1 de junio de 2026 al 30 de junio de 2026",),
    )


WITH_ROWS = _report_data(((1, "Laptop"), (2, "Mouse")))
EMPTY = _report_data(())


@pytest.mark.parametrize(
    "exporter",
    [
        CsvReportExporter(),
        JsonReportExporter(),
        XlsxReportExporter(),
        PdfReportExporter(),
        HtmlReportExporter(),
    ],
)
def test_exporters_produce_nonempty_bytes_for_rows_and_empty_data(
    exporter: ReportExporter,
) -> None:
    assert len(exporter.export(WITH_ROWS)) > 0
    assert len(exporter.export(EMPTY)) > 0


def test_csv_exporter_contains_header_and_values() -> None:
    raw = CsvReportExporter().export(WITH_ROWS).decode("utf-8")
    reader = list(csv.reader(io.StringIO(raw)))

    assert ["id", "nombre"] in reader
    assert ["1", "Laptop"] in reader
    assert CsvReportExporter().content_type == "text/csv"


def test_json_exporter_round_trips_columns_and_rows() -> None:
    raw = JsonReportExporter().export(WITH_ROWS)
    payload = json.loads(raw)

    assert payload["columns"] == ["id", "nombre"]
    assert payload["rows"] == [[1, "Laptop"], [2, "Mouse"]]
    assert payload["row_count"] == 2
    assert payload["period_label"] == "1 de junio de 2026 al 30 de junio de 2026"


def test_json_exporter_handles_empty_rows() -> None:
    payload = json.loads(JsonReportExporter().export(EMPTY))

    assert payload["rows"] == []
    assert payload["row_count"] == 0


def test_xlsx_exporter_is_parseable_and_contains_data() -> None:
    raw = XlsxReportExporter().export(WITH_ROWS)
    workbook = load_workbook(io.BytesIO(raw))
    sheet = workbook.active
    assert sheet is not None
    values = [tuple(row) for row in sheet.iter_rows(values_only=True)]

    assert ("id", "nombre") in values
    assert (1, "Laptop") in values


def test_xlsx_exporter_handles_empty_rows() -> None:
    raw = XlsxReportExporter().export(EMPTY)
    workbook = load_workbook(io.BytesIO(raw))
    sheet = workbook.active
    assert sheet is not None


def test_pdf_exporter_starts_with_pdf_signature() -> None:
    assert PdfReportExporter().export(WITH_ROWS).startswith(b"%PDF")
    assert PdfReportExporter().export(EMPTY).startswith(b"%PDF")
    assert PdfReportExporter().content_type == "application/pdf"


def test_html_exporter_escapes_and_contains_data() -> None:
    data = _report_data(((1, "A & B <C>"),))

    raw = HtmlReportExporter().export(data).decode("utf-8")

    assert "A & B <C>" not in raw
    assert "A &amp; B &lt;C&gt;" in raw
    assert "<h1>Reporte: ventas del mes pasado</h1>" in raw
    assert "<th>id</th>" in raw
    assert "<th>nombre</th>" in raw
    assert "1 de junio de 2026 al 30 de junio de 2026" in raw
    assert HtmlReportExporter().content_type == "text/html; charset=utf-8"
    assert HtmlReportExporter().file_extension == "html"


def test_html_exporter_handles_empty_rows() -> None:
    raw = HtmlReportExporter().export(EMPTY).decode("utf-8")

    assert "Sin resultados" in raw
    assert '<td colspan="2">Sin resultados</td>' in raw
